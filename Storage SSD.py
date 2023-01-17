import ctypes
import openpyxl
import datetime
from email.mime.base import MIMEBase
from email.utils import COMMASPACE, formatdate
from email import encoders

#The mail part doesn't work properly


def get_free_space_mb(folder):
    """Return folder/drive free space (in megabytes)."""
    free_bytes = ctypes.c_ulonglong(0)
    ctypes.windll.kernel32.GetDiskFreeSpaceExW(ctypes.c_wchar_p(folder), None, None, ctypes.pointer(free_bytes))
    return free_bytes.value / (1024.0 ** 2)



# Use the path to the network location as the 'folder' argument
free_space = get_free_space_mb('x: or C: ................')

# Open an existing Excel file or create a new one if it doesn't exist (use "/")
filename = 'C://where you want the file to be stored + => //storage_info.xlsx'

# Get the sheet to write the data to
sheet = None
try:
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active

    # Get the current date
    now = datetime.datetime.now()
    date = now.strftime("%Y-%m-%d %H:%M:%S")

    # Name the columns
    sheet.cell(row=1, column=1).value = "Date"
    sheet.cell(row=1, column=2).value = "Storage"

    # Find the next row
    next_row = sheet.max_row + 1

    # Write the data to the sheet
    sheet.cell(row=next_row, column=1).value = date
    sheet.cell(row=next_row, column=2).value = free_space

    # Save the workbook
    workbook.save(filename)

    print(f'Free space: {free_space/1000**2} TB')
    print(f'Data written to {filename}')

    

except IOError:
    # File does not exist, create a new workbook
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    workbook.save(filename)
    print(f'Error: {filename} does not exist.')
